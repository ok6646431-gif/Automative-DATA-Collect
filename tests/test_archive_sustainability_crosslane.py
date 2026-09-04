import hashlib
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfWriter

from orchestrator.archive_sustainability_crosslane import deduplicate_envinfo_annual_report_copies
from orchestrator.archive_user_dedup_v2 import canonicalize_user_envinfo


def write_blank_pdf(path: Path, width: float, height: float, metadata: dict[str, str]):
    path.parent.mkdir(parents=True, exist_ok=True)
    writer = PdfWriter()
    writer.add_blank_page(width=width, height=height)
    writer.add_metadata(metadata)
    with path.open('wb') as f:
        writer.write(f)


class ArchiveSustainabilityCrosslaneTests(unittest.TestCase):
    def test_envinfo_attachment_semantic_duplicate_redirects_to_official_report(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td) / '기업_환경자료'
            site = root/'01_사용자자료'/'03_환경정보공개시스템'/'광주공장'/'첨부자료'
            report_dir = root/'01_사용자자료'/'04_지속가능경영보고서'
            system = root/'90_시스템원본'/'ENVINFO'
            idx = root/'00_자료목록'
            for p in [site, report_dir, system, idx]:
                p.mkdir(parents=True, exist_ok=True)
            (idx/'README_먼저읽기.txt').write_text('Archive v2 사용 안내\n', encoding='utf-8')

            official = report_dir/'기업_지속가능경영보고서_2022.pdf'
            attachment = site/'2022_지속가능경영보고서.pdf'
            distinct = site/'2022_다른자료.pdf'
            other_year = site/'2021_지속가능경영보고서.pdf'

            write_blank_pdf(official, 595, 842, {'/Title': 'Official', '/ModDate': 'D:20220101000000'})
            write_blank_pdf(attachment, 595, 842, {'/Title': 'ENVINFO copy', '/ModDate': 'D:20221231000000'})
            write_blank_pdf(distinct, 612, 792, {'/Title': 'Different same-year PDF'})
            write_blank_pdf(other_year, 595, 842, {'/Title': 'Same rendering different year'})
            (system/'raw_2022_report.pdf').write_bytes(attachment.read_bytes())

            self.assertNotEqual(
                hashlib.sha256(official.read_bytes()).hexdigest(),
                hashlib.sha256(attachment.read_bytes()).hexdigest(),
            )

            prior = canonicalize_user_envinfo(root)
            central = root/'01_사용자자료'/'03_환경정보공개시스템'/'첨부자료_원문'
            report_copy = next(p for p in central.iterdir() if '2022_지속가능경영보고서' in p.name)
            distinct_copy = next(p for p in central.iterdir() if '2022_다른자료' in p.name)
            other_year_copy = next(p for p in central.iterdir() if '2021_지속가능경영보고서' in p.name)
            removed_rel = report_copy.relative_to(root).as_posix()
            official_rel = official.relative_to(root).as_posix()

            stats = deduplicate_envinfo_annual_report_copies(root, prior)

            self.assertEqual(stats['sustainability_crosslane_duplicate_files_removed'], 1)
            self.assertEqual(stats['sustainability_crosslane_failures'], [])
            self.assertEqual(stats['sustainability_semantic_duplicate_files_removed'], 1)
            self.assertFalse(report_copy.exists())
            self.assertTrue(official.exists())
            self.assertTrue(distinct_copy.exists(), 'Different same-year PDF must remain')
            self.assertTrue(other_year_copy.exists(), 'Same rendering from another year must remain')
            self.assertTrue((system/'raw_2022_report.pdf').exists(), 'Raw-system evidence must still exist before final system dedup')

            ref = idx/'ENVINFO_첨부자료_참조표.xlsx'
            wb = load_workbook(ref, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            headers = [str(v or '') for v in rows[0]]
            records = [dict(zip(headers, row)) for row in rows[1:]]
            redirected = [r for r in records if r.get('원래_사용자경로', '').endswith('2022_지속가능경영보고서.pdf')]
            self.assertEqual(len(redirected), 1)
            self.assertEqual(redirected[0]['최종_보존경로'], official_rel)
            self.assertIn('SEMANTIC_PDF_DUPLICATE_CANONICAL_REPORT', redirected[0]['처리'])
            self.assertNotEqual(removed_rel, official_rel)


if __name__ == '__main__':
    unittest.main()
