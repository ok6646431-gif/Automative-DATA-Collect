import csv
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from orchestrator.bat_archive import expose


def write_csv(path, rows, fields):
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', encoding='utf-8-sig', newline='') as f:
        w = csv.DictWriter(f, fieldnames=fields, extrasaction='ignore')
        w.writeheader()
        w.writerows(rows)


CANDIDATE_FIELDS = [
    'site_name', 'catalog_family', 'catalog_id', 'revision_generation', 'candidate_role', 'candidate_state', 'applicability_state',
    'publication_status', 'legal_status', 'effective_from', 'matched_industry_terms', 'matched_ksic_prefixes', 'matched_process_terms',
    'matched_utility_terms', 'evidence_channels', 'evidence_basis', 'collection_action', 'official_source_locator'
]
DOC_FIELDS = [
    'catalog_family', 'catalog_id', 'revision_id', 'revision_generation', 'publication_year', 'revision_status', 'preferred_for_matching',
    'document_part', 'volume_no', 'title', 'collection_status', 'verification_status', 'source_locator', 'source_url', 'site_names',
    'candidate_roles', 'applicability_states', 'stored_path', 'notes'
]


class BATArchiveTests(unittest.TestCase):
    def test_review_and_unpublished_candidates_create_maps_but_no_fake_pdf(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            pkg = root / 'assembled'
            archive = root / 'archive'
            pkg.mkdir()
            candidates = [
                {
                    'site_name': 'A공장', 'catalog_family': 'F_BOILER', 'catalog_id': 'BOILER_2021', 'revision_generation': 'I',
                    'candidate_role': 'COMMON_UTILITY', 'candidate_state': 'TECHNICAL_CANDIDATE', 'applicability_state': 'SUPPORTING_CANDIDATE',
                    'publication_status': 'PUBLISHED', 'legal_status': 'PUBLISHED_REFERENCE', 'matched_utility_terms': '보일러',
                    'evidence_channels': 'Management_Action_Ledger', 'evidence_basis': 'utility=보일러',
                    'collection_action': 'REVIEW_BEFORE_COLLECTION', 'official_source_locator': 'https://example.invalid/boiler'
                },
                {
                    'site_name': 'A공장', 'catalog_family': 'F_RUBBER', 'catalog_id': 'RUBBER_FUTURE', 'revision_generation': 'FUTURE',
                    'candidate_role': 'PRIMARY', 'candidate_state': 'FUTURE_PRIMARY_CANDIDATE', 'applicability_state': 'STRONG_CANDIDATE',
                    'publication_status': 'UNDER_DEVELOPMENT', 'legal_status': 'PROPOSED_FUTURE_TARGET', 'effective_from': '2029-01-01',
                    'matched_industry_terms': '타이어 및 튜브 제조업', 'evidence_channels': 'CHEM_STATS:discovery.csv:industry',
                    'evidence_basis': 'industry=타이어 및 튜브 제조업', 'collection_action': 'WAIT_FOR_PUBLICATION',
                    'official_source_locator': 'https://example.invalid/future'
                },
            ]
            docs = [
                {
                    'catalog_family': 'F_BOILER', 'catalog_id': 'BOILER_2021', 'revision_id': 'F_BOILER:I:1', 'revision_generation': 'I',
                    'publication_year': '2021', 'revision_status': 'CURRENT_PREFERRED', 'preferred_for_matching': 'true', 'document_part': '1',
                    'title': '업종공통 보일러 기준서', 'collection_status': 'REVIEW_BEFORE_COLLECTION', 'verification_status': 'NOT_DOWNLOADED',
                },
                {
                    'catalog_family': 'F_RUBBER', 'catalog_id': 'RUBBER_FUTURE', 'revision_id': 'F_RUBBER:FUTURE:1', 'revision_generation': 'FUTURE',
                    'publication_year': '2029', 'revision_status': 'FUTURE_OR_UNPUBLISHED', 'preferred_for_matching': 'true', 'document_part': '1',
                    'title': '고무제품 기준서', 'collection_status': 'NOT_YET_PUBLISHED', 'verification_status': 'NOT_APPLICABLE',
                },
            ]
            write_csv(pkg / 'BAT_Applicability_Candidates.csv', candidates, CANDIDATE_FIELDS)
            write_csv(pkg / 'output' / 'BAT_REFERENCES' / 'document_index.csv', docs, DOC_FIELDS)

            result = expose(pkg, archive)
            self.assertEqual(result['bat_archive_candidate_count'], 2)
            self.assertEqual(result['bat_archive_site_map_count'], 1)
            self.assertEqual(result['bat_archive_pdf_count'], 0)
            self.assertFalse(result['guideline_reference_present'])
            folder = archive / '02_BAT_참고자료'
            self.assertTrue((folder / '00_BAT_적용후보_및_수집현황.xlsx').exists())
            self.assertTrue((folder / '02_사업장별_적용맵' / 'A공장_BAT_적용맵.xlsx').exists())
            self.assertEqual(list((folder / '01_BAT_원문').rglob('*.pdf')), [])
            self.assertFalse((archive / '01_사용자자료' / '07_가이드라인_참고자료' / 'BAT_기준서').exists())

            wb = load_workbook(folder / '02_사업장별_적용맵' / 'A공장_BAT_적용맵.xlsx', read_only=True, data_only=True)
            self.assertEqual(wb.sheetnames, ['BAT 후보', '관련 기준서 판본'])
            self.assertEqual(wb['BAT 후보'].max_row, 3)

    def test_downloaded_documents_are_single_copy_by_family_and_revision(self):
        with tempfile.TemporaryDirectory() as td:
            root = Path(td)
            pkg = root / 'assembled'
            archive = root / 'archive'
            pkg.mkdir()
            candidates = [
                {
                    'site_name': 'A공장', 'catalog_family': 'F_TEST', 'catalog_id': 'TEST_II', 'revision_generation': 'II',
                    'candidate_role': 'PRIMARY', 'candidate_state': 'CONFIRMED', 'applicability_state': 'CONFIRMED',
                    'publication_status': 'PUBLISHED', 'legal_status': 'CURRENT', 'collection_action': 'COLLECT'
                },
                {
                    'site_name': 'B공장', 'catalog_family': 'F_TEST', 'catalog_id': 'TEST_II', 'revision_generation': 'II',
                    'candidate_role': 'PRIMARY', 'candidate_state': 'CONFIRMED', 'applicability_state': 'CONFIRMED',
                    'publication_status': 'PUBLISHED', 'legal_status': 'CURRENT', 'collection_action': 'COLLECT'
                },
            ]
            write_csv(pkg / 'BAT_Applicability_Candidates.csv', candidates, CANDIDATE_FIELDS)
            current = pkg / 'output' / 'BAT_REFERENCES' / 'files' / 'current.pdf'
            current.parent.mkdir(parents=True, exist_ok=True)
            current.write_bytes(b'%PDF-1.4\ncurrent\n%%EOF')
            old = pkg / 'output' / 'BAT_REFERENCES' / 'files' / 'old.pdf'
            old.write_bytes(b'%PDF-1.4\nold\n%%EOF')
            docs = [
                {
                    'catalog_family': 'F_TEST', 'catalog_id': 'TEST_II', 'revision_id': 'F_TEST:II:1', 'revision_generation': 'II', 'publication_year': '2022',
                    'revision_status': 'CURRENT_PREFERRED', 'preferred_for_matching': 'true', 'document_part': '1', 'title': '테스트 기준서 II',
                    'collection_status': 'DOWNLOADED', 'verification_status': 'BYTE_VERIFIED', 'stored_path': str(current.relative_to(pkg)),
                    'site_names': 'A공장;B공장', 'candidate_roles': 'PRIMARY', 'applicability_states': 'CONFIRMED'
                },
                {
                    'catalog_family': 'F_TEST', 'catalog_id': 'TEST_I', 'revision_id': 'F_TEST:I:1', 'revision_generation': 'I', 'publication_year': '2016',
                    'revision_status': 'SUPERSEDED_ARCHIVE_ONLY', 'preferred_for_matching': 'false', 'document_part': '1', 'title': '테스트 기준서 I',
                    'collection_status': 'DOWNLOADED', 'verification_status': 'BYTE_VERIFIED', 'stored_path': str(old.relative_to(pkg)),
                    'site_names': 'A공장;B공장', 'candidate_roles': 'PRIMARY', 'applicability_states': 'CONFIRMED'
                },
            ]
            write_csv(pkg / 'output' / 'BAT_REFERENCES' / 'document_index.csv', docs, DOC_FIELDS)

            result = expose(pkg, archive)
            self.assertEqual(result['bat_archive_pdf_count'], 2)
            self.assertEqual(result['bat_archive_current_pdf_count'], 1)
            self.assertEqual(result['bat_archive_superseded_pdf_count'], 1)
            self.assertEqual(result['bat_archive_site_map_count'], 2)
            self.assertEqual(result['bat_archive_family_count'], 1)
            self.assertTrue(result['guideline_reference_present'])

            folder = archive / '02_BAT_참고자료'
            family_dirs = [p for p in (folder / '01_BAT_원문').iterdir() if p.is_dir()]
            self.assertEqual(len(family_dirs), 1)
            current_files = list((family_dirs[0] / '01_현행_우선판').glob('*.pdf'))
            old_files = list((family_dirs[0] / '02_구판_아카이브').glob('*.pdf'))
            self.assertEqual(len(current_files), 1)
            self.assertEqual(len(old_files), 1)
            self.assertIn('2022_II', current_files[0].name)
            self.assertIn('2016_I', old_files[0].name)
            self.assertEqual(len(list((folder / '02_사업장별_적용맵').glob('*_BAT_적용맵.xlsx'))), 2)
            self.assertEqual(len(list(folder.rglob('*.pdf'))), 2, 'BAT PDFs must not be duplicated per site')


if __name__ == '__main__':
    unittest.main()
