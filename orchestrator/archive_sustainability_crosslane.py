"""Cross-lane semantic deduplication for annual sustainability reports.

ENV-INFO can expose a sustainability report as a downloadable attachment while the
same annual report is also collected from the company's official report catalog. The
raw PDFs can differ in document metadata while rendering the same pages. This stage
runs after ENV-INFO attachment canonicalization but before system/raw ZIP dedup, so a
proven user-facing duplicate can be removed without losing its raw-system evidence.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

import archive_user_dedup_v2 as base


USER_ROOT = base.USER_ROOT
ENVINFO_ROOT = base.ENVINFO_ROOT
SUSTAINABILITY_ROOT = base.SUSTAINABILITY_ROOT
CENTRAL_FOLDER = base.CENTRAL_FOLDER
REFERENCE_XLSX = base.REFERENCE_XLSX


def _rel(path: Path, archive_root: Path) -> str:
    return path.relative_to(archive_root).as_posix()


def _page_count(path: Path, cache: dict[Path, int]) -> int:
    if path not in cache:
        cache[path] = len(PdfReader(str(path), strict=False).pages)
    return cache[path]


def _ordered_candidate_years(path: Path, rows: list[dict]) -> list[str]:
    years: list[str] = []

    def add(value) -> None:
        year = base._year_from_name(str(value or ""))
        if year and year not in years:
            years.append(year)

    # Prefer the document's own filename year, then ENV-INFO disclosure context.
    add(path.name)
    for row in rows:
        add(row.get("공개연도"))
        add(row.get("파일명"))
        add(row.get("원래_사용자경로"))
    return years


def _reference_workbook(ref_path: Path):
    if not ref_path.exists():
        return None, None, {}, {}
    wb = load_workbook(ref_path)
    ws = wb.active
    headers = {str(cell.value or ""): idx for idx, cell in enumerate(ws[1], 1)}
    required = {"공개연도", "원래_사용자경로", "최종_보존경로", "파일명", "처리"}
    missing = sorted(required - set(headers))
    if missing:
        raise RuntimeError(f"ENVINFO attachment reference workbook missing columns: {missing}")

    rows_by_final: dict[str, list[dict]] = {}
    row_numbers_by_final: dict[str, list[int]] = {}
    for row_idx in range(2, ws.max_row + 1):
        record = {name: ws.cell(row_idx, col_idx).value for name, col_idx in headers.items()}
        final_path = str(record.get("최종_보존경로") or "")
        if not final_path:
            continue
        rows_by_final.setdefault(final_path, []).append(record)
        row_numbers_by_final.setdefault(final_path, []).append(row_idx)
    return wb, ws, rows_by_final, row_numbers_by_final


def deduplicate_envinfo_annual_report_copies(
    archive_root: str | Path,
    prior_stats: dict | None = None,
) -> dict:
    """Remove only proven ENV-INFO attachment copies of canonical annual reports.

    Candidate selection is deliberately broad: any centralized ENV-INFO PDF with a
    report year derived from its filename or reference workbook can be checked against
    the official sustainability report for that same year. A cheap page-count gate is
    applied first. Removal occurs only when the existing render-structure semantic hash
    is identical. Parse/comparison failures retain the candidate.
    """
    archive_root = Path(archive_root)
    user = archive_root / USER_ROOT
    stats = dict(prior_stats or {})

    official_dir = user / SUSTAINABILITY_ROOT
    central_dir = user / ENVINFO_ROOT / CENTRAL_FOLDER
    ref_path = archive_root / "00_자료목록" / REFERENCE_XLSX

    official_by_year: dict[str, list[Path]] = {}
    if official_dir.exists():
        for path in sorted(official_dir.glob("*.pdf")):
            if not path.is_file():
                continue
            rel = _rel(path, archive_root)
            if base._is_envinfo_generated_copy(rel):
                continue
            year = base._year_from_name(path.name)
            if year:
                official_by_year.setdefault(year, []).append(path)

    candidates = [
        path for path in sorted(central_dir.glob("*.pdf"))
        if path.is_file()
    ] if central_dir.exists() else []

    wb, ws, rows_by_final, row_numbers_by_final = _reference_workbook(ref_path)
    page_cache: dict[Path, int] = {}
    signature_cache: dict[Path, str] = {}
    candidate_years: set[str] = set()
    comparisons = 0
    page_count_checks = 0
    removed = 0
    removed_bytes = 0
    redirects: dict[str, str] = {}
    failures: list[dict] = []

    def signature(path: Path) -> str:
        if path not in signature_cache:
            signature_cache[path] = base._pdf_render_semantic_sha256(path)
        return signature_cache[path]

    for candidate in candidates:
        rel_candidate = _rel(candidate, archive_root)
        ref_rows = rows_by_final.get(rel_candidate, [])
        years = [year for year in _ordered_candidate_years(candidate, ref_rows) if year in official_by_year]
        if not years:
            continue
        candidate_years.update(years)

        matched: Path | None = None
        matched_year = ""
        candidate_sig = ""
        try:
            candidate_pages = _page_count(candidate, page_cache)
            for year in years:
                official = sorted(
                    official_by_year[year],
                    key=lambda p: base._canonical_rank(_rel(p, archive_root)),
                )
                for canonical in official:
                    page_count_checks += 1
                    if candidate_pages != _page_count(canonical, page_cache):
                        continue
                    comparisons += 1
                    candidate_sig = candidate_sig or signature(candidate)
                    if candidate_sig == signature(canonical):
                        matched = canonical
                        matched_year = year
                        break
                if matched is not None:
                    break
        except Exception as exc:
            failures.append({
                "generated_path": rel_candidate,
                "candidate_years": years,
                "error": f"{type(exc).__name__}: {exc}",
            })
            continue

        if matched is None:
            continue
        if wb is None or ws is None or not row_numbers_by_final.get(rel_candidate):
            failures.append({
                "generated_path": rel_candidate,
                "candidate_years": years,
                "error": "PROVEN_DUPLICATE_WITHOUT_ENVINFO_REFERENCE_ROW",
            })
            continue

        rel_matched = _rel(matched, archive_root)
        final_col = next(idx for idx, cell in enumerate(ws[1], 1) if str(cell.value or "") == "최종_보존경로")
        action_col = next(idx for idx, cell in enumerate(ws[1], 1) if str(cell.value or "") == "처리")
        for row_idx in row_numbers_by_final[rel_candidate]:
            ws.cell(row_idx, final_col).value = rel_matched
            action = str(ws.cell(row_idx, action_col).value or "")
            marker = "SEMANTIC_PDF_DUPLICATE_CANONICAL_REPORT"
            if marker not in action:
                ws.cell(row_idx, action_col).value = f"{action};{marker}".strip(";")

        size = candidate.stat().st_size
        candidate.unlink()
        redirects[rel_candidate] = rel_matched
        removed += 1
        removed_bytes += size

    if wb is not None and redirects:
        wb.save(ref_path)
    if wb is not None:
        wb.close()

    # Keep the original semantic counters as the aggregate contract, and expose the
    # cross-lane contribution separately for auditability.
    existing_years = set(stats.get("sustainability_semantic_candidate_years") or [])
    existing_failures = list(stats.get("sustainability_semantic_failures") or [])
    stats["sustainability_semantic_candidate_years"] = sorted(existing_years | candidate_years)
    stats["sustainability_semantic_comparisons"] = int(stats.get("sustainability_semantic_comparisons") or 0) + comparisons
    stats["sustainability_semantic_duplicate_files_removed"] = int(stats.get("sustainability_semantic_duplicate_files_removed") or 0) + removed
    stats["sustainability_semantic_duplicate_bytes_saved"] = int(stats.get("sustainability_semantic_duplicate_bytes_saved") or 0) + removed_bytes
    stats["sustainability_semantic_failures"] = existing_failures + failures
    stats.setdefault("sustainability_semantic_engine", "PYPDF_PAGE_RENDER_STRUCTURE_SHA256_V1")
    stats.update({
        "sustainability_crosslane_candidate_years": sorted(candidate_years),
        "sustainability_crosslane_page_count_checks": page_count_checks,
        "sustainability_crosslane_semantic_comparisons": comparisons,
        "sustainability_crosslane_duplicate_files_removed": removed,
        "sustainability_crosslane_duplicate_bytes_saved": removed_bytes,
        "sustainability_crosslane_redirects": redirects,
        "sustainability_crosslane_failures": failures,
    })

    readme = archive_root / "00_자료목록" / "README_먼저읽기.txt"
    if readme.exists():
        text = readme.read_text(encoding="utf-8")
        text = text.replace(
            "공식 연차 지속가능경영보고서와 같은 연도의 ENV-INFO 승격 PDF는",
            "공식 연차 지속가능경영보고서와 같은 연도의 ENV-INFO 승격 또는 첨부 PDF는",
        )
        readme.write_text(text, encoding="utf-8")

    return stats
